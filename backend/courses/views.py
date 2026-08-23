from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from .models import Course, CourseNote
from .serializers import CourseSerializer, CourseNoteSerializer
from users.models import Professor, Student
from users.permissions import user_has_role
from chatbot.models import Conversation, ChatMessage
import io
import re


def extract_text_from_attachment(attachment):
    if not attachment:
        return ""

    name = (getattr(attachment, 'name', '') or '').lower()
    data = attachment.read()
    attachment.seek(0)
    if not data:
        return ""

    if name.endswith('.txt'):
        try:
            return data.decode('utf-8')
        except UnicodeDecodeError:
            return data.decode('latin-1', errors='ignore')

    if name.endswith('.pdf'):
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            pages = [page.extract_text() or '' for page in reader.pages]
            return "\n\n".join(page for page in pages if page).strip()
        except Exception:
            return ""

    if name.endswith('.docx'):
        try:
            import docx
            document = docx.Document(io.BytesIO(data))
            paragraphs = [p.text.strip() for p in document.paragraphs if p.text and p.text.strip()]
            return "\n".join(paragraphs).strip()
        except Exception:
            return ""

    if name.endswith(('.xls', '.xlsx')):
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheet_text = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    cleaned = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if cleaned:
                        sheet_text.append(" ".join(cleaned))
            return "\n".join(sheet_text).strip()
        except Exception:
            return ""

    return ""


class CourseNoteRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CourseNote.objects.all()
    serializer_class = CourseNoteSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
            professor = getattr(self.request.user, 'professor_profile', None)
            if not professor:
                profile = getattr(self.request.user, 'profile', None)
                role_name = profile.role.nom if profile and profile.role else None
                if role_name == 'professeur':
                    professor = Professor.objects.filter(user=self.request.user).first()
            if professor:
                queryset = queryset.filter(professor=professor)
            else:
                return queryset.none()
        return queryset


class CourseRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_staff:
            secretaire = getattr(self.request.user, 'secretaire_profile', None)
            if not secretaire:
                profile = getattr(self.request.user, 'profile', None)
                if profile and getattr(profile, 'role', None) and profile.role.nom == 'secretaire_facultaire':
                    from users.models import SecretaireFacultaire
                    secretaire = SecretaireFacultaire.objects.filter(user=self.request.user).first()
            if secretaire and secretaire.faculte:
                queryset = queryset.filter(faculte=secretaire.faculte)
            else:
                return queryset.none()
        return queryset


class CourseListView(generics.ListCreateAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = super().get_queryset()
        faculte = self.request.query_params.get('faculte')
        if faculte:
            queryset = queryset.filter(faculte_id=faculte)
            
        if not self.request.user.is_staff:
            # Si c'est un secretaire, limiter à sa faculté
            profile = getattr(self.request.user, 'profile', None)
            if profile and getattr(profile, 'role', None) and profile.role.nom == 'secretaire_facultaire':
                from users.models import SecretaireFacultaire
                secretaire = SecretaireFacultaire.objects.filter(user=self.request.user).first()
                if secretaire and secretaire.faculte:
                    queryset = queryset.filter(faculte=secretaire.faculte)
            # Si c'est un professeur, limiter à ses propres cours
            elif profile and getattr(profile, 'role', None) and profile.role.nom == 'professeur':
                from users.models import Professor
                professor = Professor.objects.filter(user=self.request.user).first()
                if professor:
                    queryset = queryset.filter(professeur=professor)
        return queryset

    def perform_create(self, serializer):
        faculte = None
        is_central_admin = user_has_role(self.request.user, 'admin_central') or self.request.user.is_superuser
        if not (self.request.user.is_staff or is_central_admin):
            profile = getattr(self.request.user, 'profile', None)
            if profile and getattr(profile, 'role', None) and profile.role.nom == 'secretaire_facultaire':
                from users.models import SecretaireFacultaire
                secretaire = SecretaireFacultaire.objects.filter(user=self.request.user).first()
                if secretaire and secretaire.faculte:
                    faculte = secretaire.faculte
                else:
                    raise PermissionDenied("Vous n'êtes assigné à aucune faculté.")
            else:
                raise PermissionDenied("Seul un secrétaire facultaire ou un administrateur peut créer des cours.")
        else:
            # L'admin doit fournir la faculté dans les données
            faculte_id = self.request.data.get('faculte')
            if faculte_id:
                from users.models import Faculty
                faculte = Faculty.objects.get(code=faculte_id)

        serializer.save(faculte=faculte)


class CourseNoteListCreateView(generics.ListCreateAPIView):
    queryset = CourseNote.objects.all()
    serializer_class = CourseNoteSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    ALLOWED_EXTENSIONS = ('.pdf', '.doc', '.docx', '.xls', '.xlsx', '.txt', '.png', '.jpg', '.jpeg')

    def get_queryset(self):
        queryset = super().get_queryset()
        course_id = self.request.query_params.get('course_id')
        if course_id:
            queryset = queryset.filter(course_id=course_id)

        if not self.request.user.is_staff:
            professor = getattr(self.request.user, 'professor_profile', None)
            if not professor:
                profile = getattr(self.request.user, 'profile', None)
                role_name = profile.role.nom if profile and profile.role else None
                if role_name == 'professeur':
                    professor = Professor.objects.filter(user=self.request.user).first()
            if professor:
                queryset = queryset.filter(professor=professor)
            else:
                return queryset.none()

        return queryset

    def perform_create(self, serializer):
        professor = getattr(self.request.user, 'professor_profile', None)
        if not professor and not self.request.user.is_staff:
            profile = getattr(self.request.user, 'profile', None)
            role_name = profile.role.nom if profile and profile.role else None
            if role_name == 'professeur':
                professor = Professor.objects.filter(user=self.request.user).first()
            if not professor and role_name != 'professeur':
                raise PermissionDenied("Seul un professeur ou un administrateur peut ajouter des notes.")

        request_data = getattr(self.request, 'data', None)
        if request_data is None:
            request_data = getattr(self.request, 'POST', {})

        course = serializer.validated_data.get('course')
        if not course:
            course_id = request_data.get('course')
            if not course_id:
                raise ValidationError({'course': 'Le cours est obligatoire.'})
            course = Course.objects.filter(id=course_id).first()
            if not course:
                raise ValidationError({'course': 'Le cours est introuvable.'})

        if professor and not self.request.user.is_staff:
            if course.professeur != professor:
                raise ValidationError({'course': 'Vous ne pouvez ajouter une note que pour un cours qui vous appartient.'})

        attachment = self.request.FILES.get('attachment')
        if attachment:
            name = attachment.name.lower()
            if not any(name.endswith(ext) for ext in self.ALLOWED_EXTENSIONS):
                raise ValidationError({'attachment': 'Type de fichier non autorisé.'})

        content_value = serializer.validated_data.get('content') or ''
        serializer.save(professor=professor, content=content_value.strip())


class ProfessorStudentProgressView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        professor = getattr(request.user, 'professor_profile', None)
        if not request.user.is_staff and not professor:
            return Response({'detail': 'Accès refusé'}, status=403)

        all_students = Student.objects.filter(user__isnull=False)
        data = []

        professor_course_titles = set()
        if request.user.is_staff:
            professor_course_titles = set(course.titre for course in Course.objects.all())
        else:
            professor_course_titles.update(
                Course.objects.filter(professeur__icontains=professor.nom).values_list('titre', flat=True)
            )
            professor_course_titles.update(
                Course.objects.filter(notes__professor=professor).values_list('titre', flat=True)
            )

        for student in all_students:
            conversations = Conversation.objects.filter(user=student.user)
            if professor:
                if not professor_course_titles:
                    continue
                user_messages = ChatMessage.objects.filter(conversation__in=conversations, sender='user')
                matched = False
                for msg in user_messages:
                    text = (msg.text or '').lower()
                    if any(title.lower() in text for title in professor_course_titles):
                        matched = True
                        break
                if not matched:
                    continue

            conv_count = conversations.count()
            last_user_messages = ChatMessage.objects.filter(conversation__in=conversations, sender='user').order_by('-timestamp')[:5]
            last_message = last_user_messages.first()
            last_conversation = conversations.order_by('-updated_at').first()
            data.append({
                'student_id': student.id,
                'nom': student.nom,
                'email': student.email,
                'niveau': student.niveau,
                'conversations_count': conv_count,
                'last_message': last_message.text if last_message else None,
                'last_message_at': last_message.timestamp.isoformat() if last_message else None,
                'last_conversation_title': last_conversation.title if last_conversation else None,
                'last_conversation_updated': last_conversation.updated_at.isoformat() if last_conversation else None,
                'recent_user_queries': [msg.text for msg in last_user_messages],
            })
        return Response(data)
